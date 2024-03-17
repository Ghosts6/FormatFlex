from openpyxl import load_workbook
import xml.etree.ElementTree as ET

def excel_to_xml(excel_file, xml_file):
    
    wb = load_workbook(excel_file)
    sheet = wb.active
    
    # Create the XML structure
    root = ET.Element('data')
    
    for row in sheet.iter_rows(values_only=True):
        record = ET.SubElement(root, 'record')
        for value in row:
            ET.SubElement(record, 'item').text = str(value) if value is not None else ''
    
    tree = ET.ElementTree(root)
    
    tree.write(xml_file, encoding='utf-8', xml_declaration=True)

excel_input_file = 'input.xlsx'

xml_output_file = 'output.xml'

# Call the function to convert Excel to XML
excel_to_xml(excel_input_file, xml_output_file)
